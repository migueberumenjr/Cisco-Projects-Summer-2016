'''
Created on Jul 12, 2016

@author: miberume (Miguel Angel Berumen Jr)

@purpose: Copying and transferring data from EXS reports to templates, and extracting last dates of support for PIDs
'''
#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
import time
from time import strftime

'''
Gets the next EXS#### 
'''
exs = ''
while len(str(exs)) != 4:
    exs = input('Enter the EXS#### that you are working on (just the ####). Make sure it is 4 digits long: ')

'''
This chunk chooses the workbook and the sheet that we will extract data from
'''
analysis = openpyxl.load_workbook('C:\Users\miberume\Documents\EXS Project with Henry and Chung\EXS2900-2999\SSC EXS' + str(exs) + '.xlsx')
aSheet = analysis.active

'''
determines the column title row in the analysis sheet
'''
for begR in aSheet.columns[0]:
    if begR.value == "Item No - Up to 500".encode('utf-8'):
        titleRow = begR.row

pidCount = (aSheet.max_row - titleRow) #number of PIDS that have data to be transferred

'''
Creates a list with all the accepted country codes, taken from the Reference workbook, Country Code sheet
'''
reference = openpyxl.load_workbook('C:\Users\miberume\Documents\EXS Project with Henry and Chung\Reference for Data Upload Terms.xlsx')
rSheet = reference.get_sheet_by_name('Country Code')
refList = []

for code in rSheet.columns[0]:
    if code.row > 0 and code.row <= rSheet.max_row:
        refList.append(str(code.value))

'''
reads the column titles in the analysis sheet, and if it's a certain title (and not an empty column)
    store the values of the cells from that column in an array specific for that column
'''
pidList,serialList,ibList,quantList,lastList,extList,levelList,codeList,cityList,zipList,servList,assList,appList,condList,altList,slaList,tacList,suppList,tacCondList,combList = [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []

for colTitle in aSheet['B' + str(titleRow):get_column_letter(aSheet.max_column) + '' + str(titleRow)]:
    for titleObj in colTitle:
        if "Part" in str(titleObj.value) or "Valid" in str(titleObj.value):
            for pid in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if pid.row > titleObj.row and pid.row <= aSheet.max_row + 1:
                    pidList.append(str(pid.value))
        if "Serial" in str(titleObj.value):
            for serial in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if serial.row > titleObj.row and serial.row <= aSheet.max_row + 1:
                    serialList.append(str(serial.value))
        else:
            for i in range(0, pidCount):
                serialList.append('N/A')
        if "Contract" in str(titleObj.value):
            for ib in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if ib.row > titleObj.row and ib.row <= aSheet.max_row + 1:
                    if str(ib.value).isdigit():
                        ibList.append(str(ib.value))
                    else:
                        ibList.append("N/A")
                        print("DOUBLE CHECK THE CONTRACT NUMBER")
        if "Quantity" in str(titleObj.value):
            for quant in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if quant.row > titleObj.row and quant.row <= aSheet.max_row + 1:
                    quantList.append(quant.value)
        if "Last" in str(titleObj.value):
            for last in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if last.row > titleObj.row and last.row <= aSheet.max_row + 1:
                    if last.value == None:
                        lastList.append(str(last.value))
                    else:
                        lastList.append(last.value.strftime("%m/%d/%Y"))
        if "Extended" in str(titleObj.value):
            for ext in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if ext.row > titleObj.row and ext.row <= aSheet.max_row + 1:
                    if len(str(ext.value)) < 9:
                        extList.append(str(ext.value))
                        print("Convert the year in line " + str(ext.row) + ", column " + str(ext.column) + " to 4 digits. May be a forbidden format")
                    else:
                        extList.append(ext.value.strftime("%m/%d/%Y"))
        if "Requested Service Level" in str(titleObj.value):
            for level in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if level.row > titleObj.row and level.row <= aSheet.max_row + 1:
                    levelList.append(level.value)
        if "Country Code" in str(titleObj.value):
            for code in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if code.row > titleObj.row and code.row <= aSheet.max_row + 1:
                    if str(code.value) in refList:
                        codeList.append(str(code.value))
                    else:
                        codeList.append("N/A")
                        print("Country Code does not exist for row " + str(code.row) + ", column " + str(code.column))
        if "City" in str(titleObj.value):
            for city in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if city.row > titleObj.row and city.row <= aSheet.max_row + 1:
                    cityList.append(city.value)
        else:
            for i in range(0, pidCount):
                cityList.append('N/A')
        if "Zip" in str(titleObj.value):
            for zipCode in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if zipCode.row > titleObj.row and zipCode.row <= aSheet.max_row + 1:
                    zipList.append(zipCode.value)
        else:
            for i in range(0, pidCount):
                zipList.append('N/A')
        if "Approved Service Level" in str(titleObj.value) or "Service Level Approved" in str(titleObj.value):
            for serv in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if serv.row > titleObj.row and serv.row <= aSheet.max_row + 1:
                    if "2" in str(serv.value):
                        servList.append("2HR")
                    elif "4" in str(serv.value):
                        servList.append("4HR")
                    elif "NBD" in str(serv.value):
                        servList.append("NBD")
                    elif "RTF" in str(serv.value):
                        servList.append("RTF")
                    elif "SDS" in str(serv.value):
                        servList.append("SDS")
                    else:
                        servList.append("N/A")
        if "SSC Assessment" in str(titleObj.value):
            for ass in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if ass.row > titleObj.row and ass.row <= aSheet.max_row + 1:
                    if str(ass.value) == "Approved":
                        assList.append("Approved")
                    elif "Approved until" in str(ass.value) or "Approved for" in str(ass.value):
                        assList.append("Approved for a shorter term")
                    elif "Approved if" in str(ass.value):
                        print("ADD CONDITIONS MANUALLY FOR ROW: " + str(ass.row))
                        assList.append("Approved w/ Conditions")
                    elif "Declined - No Avail" in str(ass.value) or "Declined - No/" in str(ass.value) or "Lack" in str(ass.value) or "Declined - No Stock" in str(ass.value):
                        assList.append("Declined - No/Limited Spare Availability")
                    elif "Declined - Not" in str(ass.value) or "Not FRU" in str(ass.value) or "Invalid" in str(ass.value) or "Non HW" in str(ass.value):
                        assList.append("Declined - Not a stocked spare")
                    else:
                        print("Correct SSC Assessment Result for row " + str(ass.row) + ", column " + str(ass.column) + ". Might be an unknown format/result")
                        assList.append(str(ass.value))
        if "Extend-to" in str(titleObj.value) or "Through Date" in str(titleObj.value):
            for app in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if app.row > titleObj.row and app.row <= aSheet.max_row + 1:
                    if str(app.value) == "NA":
                        appList.append("N/A")
                    else:
                        try:
                            appList.append(app.value.strftime("%m/%d/%Y"))
                        except AttributeError:
                            appList.append(app.value)
        if "Comments" in str(titleObj.value):
            for cond in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if cond.row > titleObj.row and cond.row <= aSheet.max_row + 1:
                    condList.append(cond.value)
        if "Alternative" in str(titleObj.value):
            for alt in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if alt.row > titleObj.row and alt.row <= aSheet.max_row + 1:
                    if condList:
                        altList.append(alt.value)
                    else:
                        print("Nothing to see here")
        if "SLA" in str(titleObj.value):
            for sla in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if sla.row > titleObj.row and sla.row <= aSheet.max_row + 1:
                    slaList.append(sla.value)
        if "TAC Assessment" in str(titleObj.value):
            for tac in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if tac.row > titleObj.row and tac.row <= aSheet.max_row + 1:
                    tacList.append(tac.value)
        if "TAC Support" in str(titleObj.value):
            for supp in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if supp.row > titleObj.row and supp.row <= aSheet.max_row + 1:
                    suppList.append(supp.value)
        if "TAC Conditions" in str(titleObj.value):
            for tacCond in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if tacCond.row > titleObj.row and tacCond.row <= aSheet.max_row + 1:
                    tacCondList.append(tacCond.value)
        if "Combined" in str(titleObj.value):
            for comb in aSheet.columns[column_index_from_string(titleObj.column)-1]:
                if comb.row > titleObj.row and comb.row <= aSheet.max_row + 1:
                    combList.append(comb.value)
            
###########################################################################################################################
            
'''
This chunk chooses the workbook and sheet that we will transfer the data to
'''
template = openpyxl.load_workbook('C:\\Users\\miberume\\Documents\\EXS Project with Henry and Chung\\SSC Response.xlsx')
tSheet = template.get_sheet_by_name('MSS Assessment')

tTitleRow = 10 #Row of column titles in template is always 10
tPidCount = 0

'''       
transfers the data from each array to its corresponding column in the template sheet
'''
for tColTitle in tSheet['B' + str(tTitleRow):get_column_letter(tSheet.max_column) + '' + str(tTitleRow)]:
    for tTitleObj in tColTitle:
        if "Part" in str(tTitleObj.value) and pidList:
            tPidCount = 0
            for tPid in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tPid.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tPid.row)] = pidList[tPidCount]
                    tPidCount += 1
        if "Serial" in str(tTitleObj.value) and serialList:
            tPidCount = 0
            for tSerial in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tSerial.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tSerial.row)] = serialList[tPidCount]
                    tPidCount += 1
        if "Service Contract" in str(tTitleObj.value) and ibList:
            tPidCount = 0
            for tIB in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tIB.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tIB.row)] = ibList[tPidCount]
                    tPidCount += 1
        if "Quantity" in str(tTitleObj.value) and quantList:
            tPidCount = 0
            for tQuant in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tQuant.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tQuant.row)] = quantList[tPidCount]
                    tPidCount += 1
        if "Last" in str(tTitleObj.value) and lastList:
            tPidCount = 0
            for tLast in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tLast.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tLast.row)] = lastList[tPidCount]
                    tPidCount += 1
        if "Extended" in str(tTitleObj.value) and extList:
            tPidCount = 0
            for tExt in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tExt.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tExt.row)] = extList[tPidCount]
                    tPidCount += 1
        if "Requested Service Level" in str(tTitleObj.value) and levelList:
            tPidCount = 0
            for tLevel in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tLevel.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tLevel.row)] = levelList[tPidCount]
                    tPidCount += 1
        if "Country Code" in str(tTitleObj.value) and codeList:
            tPidCount = 0
            for tCode in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tCode.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tCode.row)] = codeList[tPidCount]
                    tPidCount += 1
        if "City" in str(tTitleObj.value) and cityList:
            tPidCount = 0
            for tCity in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tCity.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tCity.row)] = cityList[tPidCount]
                    tPidCount += 1
        if "Zip" in str(tTitleObj.value) and zipList:
            tPidCount = 0
            for tZipCode in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tZipCode.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tZipCode.row)] = zipList[tPidCount]
                    tPidCount += 1
        if "Service Level Approved" in str(tTitleObj.value) and servList:
            tPidCount = 0
            for tServ in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tServ.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tServ.row)] = servList[tPidCount]
                    tPidCount += 1
        if "SSC Assessment" in str(tTitleObj.value) and assList:
            tPidCount = 0
            for tAss in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tAss.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tAss.row)] = assList[tPidCount]
                    tPidCount += 1
        if "SSC Support Approved Through Date" in str(tTitleObj.value) and appList:
            tPidCount = 0
            for tApp in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tApp.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tApp.row)] = appList[tPidCount]
                    tPidCount += 1
        if str(tTitleObj.value) == "Conditions" and condList:
            tPidCount = 0
            for tCond in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tCond.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tCond.row)] = condList[tPidCount]
                    tPidCount += 1
        if "Alternative" in str(tTitleObj.value) and altList:
            tPidCount = 0
            for tAlt in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tAlt.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tAlt.row)] = altList[tPidCount]
                    tPidCount += 1
        if "SLA" in str(tTitleObj.value) and slaList:
            tPidCount = 0
            for tSla in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tSla.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tSla.row)] = slaList[tPidCount]
                    tPidCount += 1
        if "TAC Assessment" in str(tTitleObj.value) and tacList:
            tPidCount = 0
            for tTac in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tTac.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tTac.row)] = tacList[tPidCount]
                    tPidCount += 1
        if "TAC Support" in str(tTitleObj.value) and suppList:
            tPidCount = 0
            for tSupp in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tSupp.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tSupp.row)] = suppList[tPidCount]
                    tPidCount += 1
        if "TAC Conditions" in str(tTitleObj.value) and tacCondList:
            tPidCount = 0
            for tTacCond in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tTacCond.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tTacCond.row)] = tacCondList[tPidCount]
                    tPidCount += 1
        if "Combined" in str(tTitleObj.value) and combList:
            tPidCount = 0
            for tComb in tSheet.columns[column_index_from_string(tTitleObj.column)-1]:
                if tComb.row > tTitleRow and tPidCount < pidCount:
                    tSheet[str(tTitleObj.column) + '' + str(tComb.row)] = combList[tPidCount]
                    tPidCount += 1

###############################################################################################################

'''
This chunk gets the last date of support for the PIDs from the PID Assessment Report
    and transfers over the dates to the template
'''
if not lastList:                    
    ldos = openpyxl.load_workbook('C:\\Users\\miberume\\Documents\\EXS Project with Henry and Chung\\EXS2900-2999\\ESC EXS' + str(exs) + '.xlsx')
    
    validPID = ldos.get_sheet_by_name('Valid EoL PIDs')
    
    lastDate = "Last Date of Support"
    tPidCount2 = 0
    pidDict = {}
    
    '''
    #Determines the starting row in the PID Assessment Report
    '''
    for pidCell in validPID.columns[0]:
        if str(pidCell.value) == "Product ID":
            startRow = pidCell.row
            break
    
    '''
    Creates a dictionary for the PIDs. PIDs are the keys and LDOS dates are values. Get this from the report
    '''
    for pid in validPID.columns[0]:
        if pid.row > startRow and pid.row <= validPID.max_row:
            pidDict[str(pid.value)] = validPID[get_column_letter(33) + '' + str(pid.row)].value.strftime("%m/%d/%Y")
    '''
    #Adds the Last Dates of Support to the template excel sheet
    #This is the last programmatic data transfer. Need to add the
    #    customer name, response and expiration dates manually, 
    #    as well as the EXS number in the sheet and in the title
    '''
    for ldosCol in tSheet.columns[5]:
        if ldosCol.row > tTitleRow and tPidCount2 < pidCount and pidDict:
            if pidDict.has_key(str(tSheet['B' + str(ldosCol.row)].value)):
                ldosCol.value = pidDict.get(str(tSheet['B' + str(ldosCol.row)].value))
                tPidCount2 += 1
            else:
                print(str(pidDict.get(str(tSheet['B' + str(ldosCol.row)].value))) + " does not exist in the dictionary. Appending 'NA' for the PIDs LDoS")
                ldosCol.value = "NA"
                tPidCount2 += 1
                       
'''
#Adds series of numbers counting all the PIDs on the
#    first column
'''
n=1
for numCount in tSheet.columns[0]:
    if numCount.row > tTitleRow and numCount < (tTitleRow + len(pidList)):
        numCount.value = n
        n+=1
        
print("ALWAYS DOUBLE-CHECK THE RESULTS")

print("DOUBLE CHECK THE DATES AND ASSESSMENTS")

print("SAVE AS TEXT TAB DELIMITED AS WELL")
    
template.save('C:\Users\miberume\Documents\EXS Project with Henry and Chung\EXS2900-2999 Excel\EXS' + str(exs) + '.xlsx')
